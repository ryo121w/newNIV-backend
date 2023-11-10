import json
import os
import uuid
import openpyxl
import boto3
import tempfile
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse, HttpResponse, StreamingHttpResponse
from django.conf import settings
from django.contrib.auth import authenticate, login
from django.views.decorators.csrf import csrf_exempt
from rest_framework.views import APIView
from rest_framework.parsers import MultiPartParser
from rest_framework.response import Response
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from django.contrib.auth.models import User
from django.shortcuts import redirect
from django.contrib import messages

import io
from io import BytesIO
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.cm as cm
import numpy as np
import cloudinary
import cloudinary.uploader
import cloudinary.api
from scipy import ndimage
from sklearn.decomposition import PCA
import prince
from scipy.integrate import simps
from scipy.signal import savgol_filter
import zipstream
import zipfile
from scipy.signal import find_peaks
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
import logging


logger = logging.getLogger(__name__)
# ===============================================================================================================================


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def special_host_api(request):
    if request.user.is_superuser or request.user.is_host_approved:
        return Response({"view": "host"})
    else:
        return Response({"view": "wait"})


@csrf_exempt
def login_view(request):
    if request.method == "POST":
        data = json.loads(request.body)
        username = data.get('username')
        password = data.get('password')

        user = authenticate(username=username, password=password)

        if user is not None:
            if user.is_staff:  # is_staffフィールドをチェック
                # ログイン成功時の処理
                return JsonResponse({"message": "Login successful."}, status=200)
            else:
                # is_staffフィールドがFalseの場合
                return JsonResponse({"message": "You do not have staff access."}, status=403)
        else:
            return JsonResponse({"message": "Invalid username or password."}, status=401)
    return JsonResponse({"message": "Method not allowed."}, status=405)


@csrf_exempt
def signup_view(request):
    if request.method == "POST":
        data = json.loads(request.body)
        username = data.get('username')
        password = data.get('password')

        if User.objects.filter(username=username).exists():
            return JsonResponse({"message": "Username already exists."}, status=400)

        try:
            user = User.objects.create_user(
                username=username, password=password)
            # ユーザー登録成功時の処理をこちらに記述
            return JsonResponse({"message": "Signup successful."}, status=201)
        except Exception as e:
            return JsonResponse({"message": str(e)}, status=500)
    return JsonResponse({"message": "Method not allowed."}, status=405)


@csrf_exempt
def superuser_login(request):
    data = json.loads(request.body)
    print("Received data:", data)  # ログを追加

    username = data.get('username')
    password = data.get('password')
    user = authenticate(request, username=username, password=password)

    if user:
        print("User authenticated:", user.username)  # ユーザーが認証された場合のログ
        if user.is_superuser:
            print("User is superuser")  # ユーザーがスーパーユーザーの場合のログ
            login(request, user)
            return JsonResponse({"status": "success", "message": "Logged in as superuser."})
        else:
            print("User is not superuser")  # ユーザーがスーパーユーザーでない場合のログ
            return JsonResponse({"status": "error", "message": "You do not have superuser access."}, status=403)
    else:
        print("Authentication failed")  # 認証が失敗した場合のログ
        return JsonResponse({"status": "error", "message": "Authentication failed."}, status=401)


# ===============================================================================================================================

# ===============================================================================================================================
# Cloudinary設定
cloudinary.config(
    cloud_name=os.environ.get('CLOUDINARY_CLOUD_NAME'),
    api_key=os.environ.get('CLOUDINARY_API_KEY'),
    api_secret=os.environ.get('CLOUDINARY_API_SECRET')
)
# ===============================================================================================================================


# ===============================================================================================================================
# AWS S3設定
s3_client = boto3.client('s3',
                         aws_access_key_id=os.environ.get('AWS_ACCESS_KEY_ID'),
                         aws_secret_access_key=os.environ.get('AWS_SECRET_ACCESS_KEY'))
bucket_name = settings.AWS_STORAGE_BUCKET_NAME
# ===============================================================================================================================

# ===============================================================================================================================
# バケット内のフォルダを指定して取得することができる


def list_files(s3_client, bucket, prefix):
    response = s3_client.list_objects_v2(Bucket=bucket, Prefix=prefix)
    if 'Contents' in response:
        return [content['Key'] for content in response['Contents']]
    return []
# ===============================================================================================================================


# ===============================================================================================================================
# バケット内の特定のフォルダ内にあるデータを消去する
def delete_files_in_folder(s3_client, bucket, folder):
    """Delete all files in a specific S3 folder."""
    files = list_files(s3_client, bucket, folder)
    for file_key in files:
        s3_client.delete_object(Bucket=bucket, Key=file_key)
# ===============================================================================================================================


# ===============================================================================================================================


def upload_to_s3(file_path, s3_path):
    """
    S3にファイルをアップロードする関数
    :param file_path: アップロードするファイルのローカルパス
    :param s3_path: S3の保存先パス
    :return: None
    """
    s3_client = boto3.client('s3', aws_access_key_id=os.environ.get('AWS_ACCESS_KEY_ID'),
                             aws_secret_access_key=os.environ.get('AWS_SECRET_ACCESS_KEY'))

    bucket_name = settings.AWS_STORAGE_BUCKET_NAME

    with open(file_path, 'rb') as f:
        s3_client.upload_fileobj(f, bucket_name, s3_path)

    return
# ===============================================================================================================================

# ===============================================================================================================================
# ファイルアップロード関数


@csrf_exempt
def upload_file(request):
    if request.method == 'POST':
        try:
            print("Access Key:", os.environ.get('AWS_ACCESS_KEY_ID'))
            print("Secret Key:", os.environ.get('AWS_SECRET_ACCESS_KEY'))

            data = json.loads(request.body.decode('utf-8'))

            file_name = f"{uuid.uuid4()}.xlsx"

            workbook = openpyxl.Workbook()
            sheet = workbook.active
            for index, row in enumerate(data):
                for key, value in row.items():
                    if index == 0:
                        header_col = sheet.cell(
                            row=1, column=list(row.keys()).index(key) + 1)
                        header_col.value = key

                    cell = sheet.cell(
                        row=index + 2, column=list(row.keys()).index(key) + 1)
                    cell.value = value

            with open(file_name, 'wb') as f:
                workbook.save(f)

            s3_path = f"uploads/excel/{file_name}"

            s3_client = boto3.client('s3', aws_access_key_id=os.environ.get('AWS_ACCESS_KEY_ID'),
                                     aws_secret_access_key=os.environ.get('AWS_SECRET_ACCESS_KEY'))

            existing_files = list_files(
                s3_client, settings.AWS_STORAGE_BUCKET_NAME, 'uploads/excel/')
            for file_key in existing_files:
                s3_client.delete_object(
                    Bucket=settings.AWS_STORAGE_BUCKET_NAME, Key=file_key)

            s3_client.upload_file(
                file_name, settings.AWS_STORAGE_BUCKET_NAME, s3_path)

            os.remove(file_name)

            file_url = f"https://{settings.AWS_STORAGE_BUCKET_NAME}.s3.amazonaws.com/{s3_path}"

            return JsonResponse({'message': 'Data processed and saved to S3 successfully!', 'file_url': file_url})
        except json.JSONDecodeError:
            return JsonResponse({'message': 'Failed to decode JSON data.'}, status=400)

    return JsonResponse({'message': 'Only POST requests are allowed.'})
# ===============================================================================================================================


# ===============================================================================================================================
# NIRスペクトル関数


def generate_spectrum_graph(request):
    # Cloudinaryの設定
    cloudinary.config(
        cloud_name=os.environ.get('CLOUDINARY_CLOUD_NAME'),
        api_key=os.environ.get('CLOUDINARY_API_KEY'),
        api_secret=os.environ.get('CLOUDINARY_API_SECRET')
    )

    s3_client = boto3.client('s3', aws_access_key_id=os.environ.get('AWS_ACCESS_KEY_ID'),
                             aws_secret_access_key=os.environ.get('AWS_SECRET_ACCESS_KEY'))
    bucket_name = settings.AWS_STORAGE_BUCKET_NAME

    # S3内の'uploads/excel/'ディレクトリから最新のファイルを取得
    # (list_files関数の定義がないため、この部分を具体的には確認できません。)
    files = list_files(s3_client, bucket_name, 'uploads/excel/')
    if not files:
        return HttpResponse('No files found in S3 bucket.')

    latest_uploaded_file = sorted(files)[-1]

    # S3からExcelファイルをダウンロード
    obj = s3_client.get_object(Bucket=bucket_name, Key=latest_uploaded_file)
    df = pd.read_excel(BytesIO(obj['Body'].read()))
    df = df[(df['波長'] >= 6000) & (df['波長'] <= 8000)]

    # グラフ生成
    plt.figure(figsize=(10, 6))
    plt.xlim(6000, 8000)

    # 6000-8000の範囲での最大値を検知し、y軸の上限を設定
    max_y_val_within_range = df[df.columns[1:]].max().max()
    plt.ylim(0, max_y_val_within_range + 0.1)

    concentrations = None
    concentrations_columns = concentrations if concentrations else list(
        df.columns[1:])
    colors = plt.cm.rainbow(np.linspace(0, 0.5, len(concentrations_columns)))

    for col_name, color in zip(concentrations_columns, colors):
        if col_name in df.columns:
            plt.plot(df['波長'], df[col_name], label=col_name, color=color)

    plt.title('NIR Spectrum')
    plt.xlabel('Wavelength (cm-1)')
    plt.ylabel('Absorbance')
    plt.legend()

    # グラフをバイナリのIOストリームとして保存
    img_data = io.BytesIO()
    plt.savefig(img_data, format='png')
    img_data.seek(0)
    plt.close()

    # Cloudinaryに保存されている古いイメージを削除
    folder_name = 'spectrums'
    stored_images = cloudinary.api.resources(
        type='upload', prefix=folder_name, max_results=500)
    for image in stored_images['resources']:
        cloudinary.uploader.destroy(image['public_id'])

    # 新しいイメージをCloudinaryのフォルダにアップロード
    upload_response = cloudinary.uploader.upload(
        img_data,
        folder=folder_name,
        use_filename=True,
        unique_filename=False
    )
    graph_url = upload_response['url']

    return HttpResponse(graph_url)
# ===============================================================================================================================


# ===============================================================================================================================
# モル吸光係数
# モル濃度の情報を取得


def get_molarities_from_excel(file_path):
    with tempfile.NamedTemporaryFile(delete=False) as temp_file:
        s3_client = boto3.client('s3', aws_access_key_id=os.environ.get('AWS_ACCESS_KEY_ID'),
                                 aws_secret_access_key=os.environ.get('AWS_SECRET_ACCESS_KEY'))
    bucket_name = settings.AWS_STORAGE_BUCKET_NAME
    s3_client.download_file(bucket_name, file_path, temp_file.name)
    df = pd.read_excel(temp_file.name)
    molarities = df.columns[1:].tolist()
    return molarities

# ファイルからモル濃度の情報を取得


def get_files_from_s3(request):
    s3_client = boto3.client('s3')
    bucket_name = settings.AWS_STORAGE_BUCKET_NAME
    files = list_files(s3_client, bucket_name, 'uploads/excel/')

    if not files:
        return JsonResponse({'error': 'No files found in S3 bucket.'}, status=404)

    # エクセルファイルから濃度情報を取得 (ここでは最初のファイルを仮で利用)
    file_path = files[0]
    molarities = get_molarities_from_excel(file_path)
    return JsonResponse({'files': files, 'molarities': molarities})

# エクセルファイルダウンロード


def download_excel(request):
    s3_client = boto3.client('s3',
                             aws_access_key_id=os.environ.get(
                                 'AWS_ACCESS_KEY_ID'),
                             aws_secret_access_key=os.environ.get('AWS_SECRET_ACCESS_KEY'))

    # S3内の'uploads/excel/'ディレクトリから最新のファイルを取得（あなたのコードを前提としています）
    bucket_name = 'newniv-bucket'
    files = list_files(s3_client, bucket_name, 'processed_data')

    if not files:
        return HttpResponse('No files found in S3 bucket.')

    try:
        # 最新のファイルのキー
        latest_file_key = files[0]

        # メモリ上のバイナリストリームとしてファイルを取得
        file_stream = s3_client.get_object(
            Bucket=bucket_name, Key=latest_file_key)['Body']

        # クライアントに送信するためのレスポンスを作成
        response = HttpResponse(file_stream.read(),
                                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        # 最新のファイル名をそのまま使用
        response[
            'Content-Disposition'] = f'attachment; filename="{latest_file_key.split("/")[-1]}"'

        return response

    except Exception as e:
        print(e)
        return HttpResponse("An error occurred while downloading the file.")

# AWSのファイルをリストで取得


def list_files(s3_client, bucket_name, prefix):
    # バケットから特定のプレフィックスを持つオブジェクトのリストを取得
    response = s3_client.list_objects_v2(Bucket=bucket_name, Prefix=prefix)
    return [content['Key'] for content in response.get('Contents', [])]
# モル吸光係数


class ConcentrationGraphView(APIView):
    parser_classes = (MultiPartParser,)

    def post(self, request):
        print(f"Debug: Received POST data: {request.data}")
        concentrations = request.data.getlist('concentrations[]', [])
        print(f"Debug: Received concentrations: {concentrations}")

        # S3 configuration
        s3_client = boto3.client('s3', aws_access_key_id=os.environ.get('AWS_ACCESS_KEY_ID'),
                                 aws_secret_access_key=os.environ.get('AWS_SECRET_ACCESS_KEY'))
        bucket_name = os.environ.get('AWS_STORAGE_BUCKET_NAME')

        # S3内の'uploads/excel/'ディレクトリから最新のファイルを取得
        files = list_files(s3_client, bucket_name, 'uploads/excel/')
        if not files:
            return HttpResponse('No files found in S3 bucket.')

        latest_file_key = files[-1]
        local_path = "/tmp/latest_file.xlsx"
        s3_client.download_file(bucket_name, latest_file_key, local_path)

        df = pd.read_excel(local_path)
        columns = df.columns.drop('波長')
        print(f"Debug: Excel columns: {columns.tolist()}")

        if len(columns) != len(concentrations):
            error_message = f'Mismatch between number of data columns ({len(columns)}) and provided concentrations ({len(concentrations)}). Columns: {columns.tolist()}, Concentrations: {concentrations}'
            return Response({'error': error_message}, status=status.HTTP_400_BAD_REQUEST)

        # Create a new dataframe for normalized data
        normalized_df = pd.DataFrame()
        normalized_df['波長'] = df['波長']

        plt.figure(figsize=(10, 6))
        plt.xlim(8000, 6000)

        max_val = 0  # Initialize max_val to be updated for each processed column

        colors = cm.rainbow(np.linspace(0, 0.5, len(columns)))

        for i, (column, color) in enumerate(zip(columns, colors)):
            # Process/normalize the column data here
            norm_column = df[column] / float(concentrations[i])

            # Store normalized data in the new dataframe
            normalized_df[column] = norm_column

            # Update max_val if new max found
            current_max = norm_column[(df['波長'] >= 6000) & (
                df['波長'] <= 8000)].max()
            max_val = max(max_val, current_max)

            plt.plot(df['波長'], norm_column,
                     label=f'{column} - {concentrations[i]}M', color=color)

        plt.ylim(0, max_val + 0.01)  # Set y limit based on max_val

        plt.title('NIR Spectrum of LiCl with Concentrations')
        plt.xlabel('Wavelength (cm-1)')
        plt.ylabel('Absorbance')
        plt.legend()

        graph_filename = 'concentration_nir_spectrum.png'
        graph_dir = 'static'
        graph_filepath = os.path.join(graph_dir, graph_filename)

        if not os.path.exists(graph_dir):
            os.makedirs(graph_dir)

        plt.savefig(graph_filepath)
        plt.close()

        # Delete previous data in the processed_data folder
        delete_files_in_folder(s3_client, bucket_name, 'processed_data/')

        # Save the normalized data to a new S3 folder
        processed_excel_path = os.path.join(
            graph_dir, 'processed_data_normalized.xlsx')
        normalized_df.to_excel(processed_excel_path, index=False)
        s3_upload_normalized_path = f'processed_data/processed_data_normalized.xlsx'
        s3_client.upload_file(processed_excel_path,
                              bucket_name, s3_upload_normalized_path)

        # Cloudinaryの設定
        cloudinary.config(
            cloud_name=os.environ.get('CLOUDINARY_CLOUD_NAME'),
            api_key=os.environ.get('CLOUDINARY_API_KEY'),
            api_secret=os.environ.get('CLOUDINARY_API_SECRET')
        )

        # Cloudinaryに保存されている古いイメージを削除
        folder_name = 'concentration'
        stored_images = cloudinary.api.resources(
            type='upload', prefix=f"{folder_name}/", max_results=500)
        for image in stored_images['resources']:
            cloudinary.uploader.destroy(image['public_id'])

        # グラフをCloudinaryのフォルダにアップロード
        upload_response = cloudinary.uploader.upload(
            graph_filepath, folder=folder_name, use_filename=True, unique_filename=False)
        cloudinary_url = upload_response['url']

        response_data = {'graph_url': cloudinary_url}
        return JsonResponse(response_data)
# ===============================================================================================================================


# ===============================================================================================================================
# 二次微分
class SecondDerivativeGraphView(APIView):

    def post(self, request):
        s3_client = boto3.client('s3')
        bucket_name = 'newniv-bucket'

        files = list_files(s3_client, bucket_name, 'processed_data/')
        if not files:
            return HttpResponse('No files found in S3 bucket.')

        latest_file_key = files[-1]
        local_path = "/tmp/latest_file.xlsx"
        s3_client.download_file(bucket_name, latest_file_key, local_path)

        df = pd.read_excel(local_path)
        columns = df.columns.drop('波長')

        # Create a copy of the dataframe to store the second derivative data
        derivative_df = df.copy()

        plt.figure(figsize=(10, 6))
        plt.xlim(8000, 6000)

        # Initialize variables to find max and min values dynamically
        max_val, min_val = None, None

        colors = plt.cm.rainbow(np.linspace(0, 1, len(columns)))

        for col, c in zip(columns, colors):
            if col.startswith('Molar_Absorptivity_'):
                continue

            smoothed_data = ndimage.gaussian_filter1d(df[col], sigma=10)
            y = ndimage.gaussian_filter1d(smoothed_data, sigma=10, order=2)
            derivative_df[col] = y

            # Dynamically find max and min values in the specified range
            current_max = y[(df['波長'] >= 6000) & (df['波長'] <= 8000)].max()
            current_min = y[(df['波長'] >= 6000) & (df['波長'] <= 8000)].min()

            # Update max and min values if new extremes found
            max_val = current_max if max_val is None else max(
                max_val, current_max)
            min_val = current_min if min_val is None else min(
                min_val, current_min)

            plt.plot(df['波長'], y, label=col, color=c)

        # Set y limit based on dynamically found max and min values
        plt.ylim(min_val, max_val)

        plt.title('Second Derivative of NIR Spectrum')
        plt.xlabel('Wavelength (cm-1)')
        plt.ylabel('Second Derivative of Absorbance')
        plt.legend(loc='upper right')

        graph_filename = 'second_derivative_nir_spectrum.png'
        graph_dir = 'static'
        graph_filepath = os.path.join(graph_dir, graph_filename)

        if not os.path.exists(graph_dir):
            os.makedirs(graph_dir)

        plt.savefig(graph_filepath)
        plt.close()

        # Save the second derivative data to S3
        processed_excel_path = os.path.join(
            graph_dir, 'second_derivative_data.xlsx')
        derivative_df.to_excel(processed_excel_path, index=False)
        s3_upload_path = f'second_derivative/second_derivative_data.xlsx'
        s3_client.upload_file(processed_excel_path,
                              bucket_name, s3_upload_path)

        # Cloudinaryの設定
        cloudinary.config(
            cloud_name=os.environ.get('CLOUDINARY_CLOUD_NAME'),
            api_key=os.environ.get('CLOUDINARY_API_KEY'),
            api_secret=os.environ.get('CLOUDINARY_API_SECRET')
        )

        # Cloudinaryに保存されている古いイメージを削除
        folder_name = 'SecondDerivative'
        stored_images = cloudinary.api.resources(
            type='upload', prefix=f"{folder_name}/", max_results=500)
        for image in stored_images['resources']:
            cloudinary.uploader.destroy(image['public_id'])

        # グラフをCloudinaryのフォルダにアップロード
        upload_response = cloudinary.uploader.upload(
            graph_filepath, folder=folder_name, use_filename=True, unique_filename=False)
        cloudinary_url = upload_response['url']

        response_data = {'graph_url': cloudinary_url}
        return JsonResponse(response_data)
# 二次微分のエクセルデータダウンロード


def second_derivative_download(request):
    s3_client = boto3.client('s3',
                             aws_access_key_id=os.environ.get(
                                 'AWS_ACCESS_KEY_ID'),
                             aws_secret_access_key=os.environ.get('AWS_SECRET_ACCESS_KEY'))

    # S3内の'second_derivative/'ディレクトリから最新のファイルを取得
    bucket_name = 'newniv-bucket'
    prefix = 'second_derivative/'
    files = list_files(s3_client, bucket_name, prefix)

    if not files:
        return HttpResponse('No files found in S3 bucket under the specified prefix.')

    try:
        # 最新のファイルのキー
        latest_file_key = files[-1]  # 最新のファイルを取得するために[-1]を使用

        # メモリ上のバイナリストリームとしてファイルを取得
        file_stream = s3_client.get_object(
            Bucket=bucket_name, Key=latest_file_key)['Body']

        # クライアントに送信するためのレスポンスを作成
        response = HttpResponse(file_stream.read(),
                                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        # 最新のファイル名をそのまま使用
        response[
            'Content-Disposition'] = f'attachment; filename="{latest_file_key.split("/")[-1]}"'

        return response

    except Exception as e:
        print(e)
        return HttpResponse("An error occurred while downloading the file.")
# ===============================================================================================================================


# ===============================================================================================================================
# 三次微分
class ThirdDerivativeGraphView(APIView):

    def post(self, request):
        s3_client = boto3.client('s3')
        bucket_name = 'newniv-bucket'

        files = list_files(s3_client, bucket_name, 'processed_data/')
        if not files:
            return HttpResponse('No files found in S3 bucket.')

        latest_file_key = files[-1]
        local_path = "/tmp/latest_file.xlsx"
        s3_client.download_file(bucket_name, latest_file_key, local_path)

        df = pd.read_excel(local_path)
        columns = df.columns.drop('波長')

        # Create a copy of the dataframe to store the third derivative data
        derivative_df = df.copy()

        plt.figure(figsize=(10, 6))
        plt.xlim(8000, 6000)

        # Initialize variables to find max and min values dynamically
        max_val, min_val = None, None

        colors = plt.cm.rainbow(np.linspace(0, 1, len(columns)))

        for col, c in zip(columns, colors):
            if col.startswith('Molar_Absorptivity_'):
                continue

            smoothed_data = ndimage.gaussian_filter1d(df[col], sigma=10)
            y = ndimage.gaussian_filter1d(smoothed_data, sigma=10, order=3)
            derivative_df[col] = y

            # Dynamically find max and min values in the specified range
            current_max = y[(df['波長'] >= 6000) & (df['波長'] <= 8000)].max()
            current_min = y[(df['波長'] >= 6000) & (df['波長'] <= 8000)].min()

            # Update max and min values if new extremes found
            max_val = current_max if max_val is None else max(
                max_val, current_max)
            min_val = current_min if min_val is None else min(
                min_val, current_min)

            plt.plot(df['波長'], y, label=col, color=c)

        # Set y limit based on dynamically found max and min values
        plt.ylim(min_val, max_val)

        plt.title('Third Derivative of NIR Spectrum')
        plt.xlabel('Wavelength (cm-1)')
        plt.ylabel('Third Derivative of Absorbance')
        plt.legend(loc='upper right')

        graph_filename = 'third_derivative_nir_spectrum.png'
        graph_dir = 'static'
        graph_filepath = os.path.join(graph_dir, graph_filename)

        if not os.path.exists(graph_dir):
            os.makedirs(graph_dir)

        plt.savefig(graph_filepath)
        plt.close()

        # Save the third derivative data to S3
        processed_excel_path = os.path.join(
            graph_dir, 'third_derivative_data.xlsx')
        derivative_df.to_excel(processed_excel_path, index=False)
        s3_upload_path = f'third_derivative/third_derivative_data.xlsx'
        s3_client.upload_file(processed_excel_path,
                              bucket_name, s3_upload_path)

        # Cloudinaryの設定
        cloudinary.config(
            cloud_name=os.environ.get('CLOUDINARY_CLOUD_NAME'),
            api_key=os.environ.get('CLOUDINARY_API_KEY'),
            api_secret=os.environ.get('CLOUDINARY_API_SECRET')
        )

        # Cloudinaryに保存されている古いイメージを削除
        folder_name = 'ThirdDerivative'
        stored_images = cloudinary.api.resources(
            type='upload', prefix=f"{folder_name}/", max_results=500)
        for image in stored_images['resources']:
            cloudinary.uploader.destroy(image['public_id'])

        # グラフをCloudinaryのフォルダにアップロード
        upload_response = cloudinary.uploader.upload(
            graph_filepath, folder=folder_name, use_filename=True, unique_filename=False)
        cloudinary_url = upload_response['url']

        response_data = {'graph_url': cloudinary_url}
        return JsonResponse(response_data)
# 三次微分のエクセルデータダウンロード


def third_derivative_download(request):
    s3_client = boto3.client('s3',
                             aws_access_key_id=os.environ.get(
                                 'AWS_ACCESS_KEY_ID'),
                             aws_secret_access_key=os.environ.get('AWS_SECRET_ACCESS_KEY'))

    # S3内の'second_derivative/'ディレクトリから最新のファイルを取得
    bucket_name = 'newniv-bucket'
    prefix = 'third_derivative/'
    files = list_files(s3_client, bucket_name, prefix)

    if not files:
        return HttpResponse('No files found in S3 bucket under the specified prefix.')

    try:
        # 最新のファイルのキー
        latest_file_key = files[-1]  # 最新のファイルを取得するために[-1]を使用

        # メモリ上のバイナリストリームとしてファイルを取得
        file_stream = s3_client.get_object(
            Bucket=bucket_name, Key=latest_file_key)['Body']

        # クライアントに送信するためのレスポンスを作成
        response = HttpResponse(file_stream.read(),
                                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        # 最新のファイル名をそのまま使用
        response[
            'Content-Disposition'] = f'attachment; filename="{latest_file_key.split("/")[-1]}"'

        return response

    except Exception as e:
        print(e)
        return HttpResponse("An error occurred while downloading the file.")
# ===============================================================================================================================


# ===============================================================================================================================

# 四次微分
class FourthDerivativeGraphView(APIView):

    def post(self, request):
        s3_client = boto3.client('s3')
        bucket_name = 'newniv-bucket'

        files = list_files(s3_client, bucket_name, 'processed_data/')
        if not files:
            return HttpResponse('No files found in S3 bucket.')

        latest_file_key = files[-1]
        local_path = "/tmp/latest_file.xlsx"
        s3_client.download_file(bucket_name, latest_file_key, local_path)

        df = pd.read_excel(local_path)
        columns = df.columns.drop('波長')

        # Create a copy of the dataframe to store the fourth derivative data
        derivative_df = df.copy()

        plt.figure(figsize=(10, 6))
        plt.xlim(8000, 6000)

        # Initialize variables to find max and min values dynamically
        max_val, min_val = None, None

        colors = plt.cm.rainbow(np.linspace(0, 1, len(columns)))

        for col, c in zip(columns, colors):
            if col.startswith('Molar_Absorptivity_'):
                continue

            smoothed_data = ndimage.gaussian_filter1d(df[col], sigma=10)
            y = ndimage.gaussian_filter1d(smoothed_data, sigma=10, order=4)
            derivative_df[col] = y

            # Dynamically find max and min values in the specified range
            current_max = y[(df['波長'] >= 6000) & (df['波長'] <= 8000)].max()
            current_min = y[(df['波長'] >= 6000) & (df['波長'] <= 8000)].min()

            # Update max and min values if new extremes found
            max_val = current_max if max_val is None else max(
                max_val, current_max)
            min_val = current_min if min_val is None else min(
                min_val, current_min)

            plt.plot(df['波長'], y, label=col, color=c)

        # Set y limit based on dynamically found max and min values
        plt.ylim(min_val, max_val)

        plt.title('Fourth Derivative of NIR Spectrum')
        plt.xlabel('Wavelength (cm-1)')
        plt.ylabel('Fourth Derivative of Absorbance')
        plt.legend(loc='upper right')

        graph_filename = 'fourth_derivative_nir_spectrum.png'
        graph_dir = 'static'
        graph_filepath = os.path.join(graph_dir, graph_filename)

        if not os.path.exists(graph_dir):
            os.makedirs(graph_dir)

        plt.savefig(graph_filepath)
        plt.close()

        # Save the fourth derivative data to S3
        processed_excel_path = os.path.join(
            graph_dir, 'fourth_derivative_data.xlsx')
        derivative_df.to_excel(processed_excel_path, index=False)
        s3_upload_path = f'fourth_derivative/fourth_derivative_data.xlsx'
        s3_client.upload_file(processed_excel_path,
                              bucket_name, s3_upload_path)

        # Cloudinary configuration
        cloudinary.config(
            cloud_name=os.environ.get('CLOUDINARY_CLOUD_NAME'),
            api_key=os.environ.get('CLOUDINARY_API_KEY'),
            api_secret=os.environ.get('CLOUDINARY_API_SECRET')
        )

        # Delete old images stored in Cloudinary
        folder_name = 'FourthDerivative'
        stored_images = cloudinary.api.resources(
            type='upload', prefix=f"{folder_name}/", max_results=500)
        for image in stored_images['resources']:
            cloudinary.uploader.destroy(image['public_id'])

        # Upload the graph to Cloudinary folder
        upload_response = cloudinary.uploader.upload(
            graph_filepath, folder=folder_name, use_filename=True, unique_filename=False)
        cloudinary_url = upload_response['url']

        response_data = {'graph_url': cloudinary_url}
        return JsonResponse(response_data)
# 四次微分のエクセルデータダウンロード


def fourth_derivative_download(request):
    s3_client = boto3.client('s3',
                             aws_access_key_id=os.environ.get(
                                 'AWS_ACCESS_KEY_ID'),
                             aws_secret_access_key=os.environ.get('AWS_SECRET_ACCESS_KEY'))

    # S3内の'second_derivative/'ディレクトリから最新のファイルを取得
    bucket_name = 'newniv-bucket'
    prefix = 'fourth_derivative/'
    files = list_files(s3_client, bucket_name, prefix)

    if not files:
        return HttpResponse('No files found in S3 bucket under the specified prefix.')

    try:
        # 最新のファイルのキー
        latest_file_key = files[-1]  # 最新のファイルを取得するために[-1]を使用

        # メモリ上のバイナリストリームとしてファイルを取得
        file_stream = s3_client.get_object(
            Bucket=bucket_name, Key=latest_file_key)['Body']

        # クライアントに送信するためのレスポンスを作成
        response = HttpResponse(file_stream.read(),
                                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        # 最新のファイル名をそのまま使用
        response[
            'Content-Disposition'] = f'attachment; filename="{latest_file_key.split("/")[-1]}"'

        return response

    except Exception as e:
        print(e)
        return HttpResponse("An error occurred while downloading the file.")
# ===============================================================================================================================


# ===============================================================================================================================
# 差スペクトル
class DifferenceGraphView(APIView):
    parser_classes = [MultiPartParser]

    s3 = boto3.client('s3')
    BUCKET_NAME = 'newniv-bucket'

    def list_files_in_s3(self, s3_client, bucket_name, prefix):
        response = s3_client.list_objects_v2(Bucket=bucket_name, Prefix=prefix)
        if 'Contents' not in response:
            return []
        files = [item['Key'] for item in response['Contents']]
        return files

    def fetch_latest_data_from_s3(self):
        s3_client = boto3.client('s3')
        bucket_name = 'newniv-bucket'

        files = self.list_files_in_s3(
            s3_client, bucket_name, 'processed_data/')
        if not files:
            raise Exception('No files found in S3 bucket.')

        latest_file_key = files[-1]
        local_path = "/tmp/latest_file.xlsx"
        s3_client.download_file(bucket_name, latest_file_key, local_path)

        df = pd.read_excel(local_path)
        if 'Unnamed: 0' in df.columns:
            df = df.drop(columns='Unnamed: 0')

        return df

    def post(self, request, *args, **kwargs):
        df = self.fetch_latest_data_from_s3()
        if df is None:
            return Response({"error": "Could not fetch the latest data from S3"}, status=status.HTTP_400_BAD_REQUEST)

        zero_m_data = df.get('0M')
        if zero_m_data is None:
            return Response({"error": "0M column not found in the data"}, status=status.HTTP_400_BAD_REQUEST)

        columns = [col for col in df.columns if col not in [
            '0M', '波長'] and not col.startswith('Molar_Absorptivity_')]

        for column in columns:
            df[column] -= zero_m_data
            baseline = np.polyfit(df['波長'], df[column], 3)
            baseline = np.polyval(baseline, df['波長'])
            df[column] -= baseline

        df = df.drop(columns='0M')

        plt.figure(figsize=(10, 6))
        plt.xlim(8000, 6000)

        # Initialize variables to find max and min values dynamically
        max_val, min_val = None, None

        colors = cm.rainbow(np.linspace(0, 0.5, len(columns)))
        for column, color in zip(columns, colors):
            # Dynamically find max and min values in the specified range
            current_max = df[column][(df['波長'] >= 6000)
                                     & (df['波長'] <= 8000)].max()
            current_min = df[column][(df['波長'] >= 6000)
                                     & (df['波長'] <= 8000)].min()

            # Update max and min values if new extremes found
            max_val = current_max if max_val is None else max(
                max_val, current_max)
            min_val = current_min if min_val is None else min(
                min_val, current_min)

            plt.plot(df['波長'], df[column], label=column, color=color)

        # Set y limit based on dynamically found max and min values
        plt.ylim(min_val, max_val)

        plt.title('Difference Spectrum with Baseline Correction')
        plt.xlabel('Wavelength (cm-1)')
        plt.ylabel('Difference Intensity')
        plt.legend()

        image_path = "/tmp/difference_graph_corrected.png"
        plt.savefig(image_path)
        plt.close()
        cloudinary_upload = cloudinary.uploader.upload(
            image_path, folder="Difference", public_id="difference_graph_corrected")

        # 不要な最初の行 (行A) を削除
        df = df.iloc[1:].reset_index(drop=True)

        difference_data_path = "/tmp/difference_data.xlsx"
        df.to_excel(difference_data_path, index=False)  # 無条件で最初の行を削除
        self.s3.upload_file(difference_data_path,
                            self.BUCKET_NAME, 'difference/difference_data.xlsx')

        image_url = cloudinary_upload['secure_url']
        return JsonResponse({"graph_url": image_url})
# 差スペクトルのデータをダウンロード


def difference_download(request):
    s3_client = boto3.client('s3',
                             aws_access_key_id=os.environ.get(
                                 'AWS_ACCESS_KEY_ID'),
                             aws_secret_access_key=os.environ.get('AWS_SECRET_ACCESS_KEY'))

    # S3内の'second_derivative/'ディレクトリから最新のファイルを取得
    bucket_name = 'newniv-bucket'
    prefix = 'difference/'
    files = list_files(s3_client, bucket_name, prefix)

    if not files:
        return HttpResponse('No files found in S3 bucket under the specified prefix.')

    try:
        # 最新のファイルのキー
        latest_file_key = files[-1]  # 最新のファイルを取得するために[-1]を使用

        # メモリ上のバイナリストリームとしてファイルを取得
        file_stream = s3_client.get_object(
            Bucket=bucket_name, Key=latest_file_key)['Body']

        # クライアントに送信するためのレスポンスを作成
        response = HttpResponse(file_stream.read(),
                                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        # 最新のファイル名をそのまま使用
        response[
            'Content-Disposition'] = f'attachment; filename="{latest_file_key.split("/")[-1]}"'

        return response

    except Exception as e:
        print(e)
        return HttpResponse("An error occurred while downloading the file.")
# ===============================================================================================================================

# PCA実行


# ===============================================================================================================================

class PrincipalComponentAnalysisView(APIView):

    def post(self, request):
        s3_client = boto3.client('s3')
        bucket_name = 'newniv-bucket'

        files = list_files(s3_client, bucket_name, 'processed_data/')
        if not files:
            return Response({'error': 'No files found in S3 bucket.'}, status=404)

        latest_file_key = files[-1]
        local_path = "/tmp/latest_file.xlsx"
        s3_client.download_file(bucket_name, latest_file_key, local_path)

        df = pd.read_excel(local_path)

        # '波長'の列を除外
        pca_data = df.drop(columns=['波長'])

        # PCAのコンポーネント数を動的に設定
        n_components = min(pca_data.shape[0], pca_data.shape[1])
        pca = PCA(n_components=n_components)
        pca_result = pca.fit_transform(pca_data)

        plt.figure(figsize=(10, 6))
        plt.plot(range(1, n_components+1),
                 pca.explained_variance_ratio_.cumsum(), marker='o', linestyle='--')
        plt.title('Explained Variance by Components')
        plt.xlabel('Number of Components')
        plt.ylabel('Cumulative Explained Variance')
        plt.savefig("/tmp/pca_scree_plot.png")

        # Save the PCA result data to S3
        pca_result_df = pd.DataFrame(data=pca_result, columns=[
                                     f'PC{i}' for i in range(1, n_components+1)])
        pca_result_df.to_excel("/tmp/pca_result.xlsx", index=False)
        s3_client.upload_file("/tmp/pca_result.xlsx",
                              bucket_name, "principal_analysis/pca_result.xlsx")

        # Configure Cloudinary
        cloudinary.config(
            cloud_name=os.environ.get('CLOUDINARY_CLOUD_NAME'),
            api_key=os.environ.get('CLOUDINARY_API_KEY'),
            api_secret=os.environ.get('CLOUDINARY_API_SECRET')
        )

        # Upload image to Cloudinary
        upload_response = cloudinary.uploader.upload(
            "/tmp/pca_scree_plot.png",
            folder="principal_analysis",
            use_filename=True,
            unique_filename=False
        )

        return Response({"graph_url": upload_response['url']}, status=200)
# ===============================================================================================================================

# MCA実行


# ===============================================================================================================================

class MCAnalysis(APIView):
    def post(self, request):
        # S3からのデータ取得コード
        s3_client = boto3.client('s3')
        bucket_name = 'newniv-bucket'

        files = list_files(s3_client, bucket_name, 'processed_data/')
        if not files:
            return Response({'error': 'No files found in S3 bucket.'}, status=404)

        latest_file_key = files[-1]
        local_path = "/tmp/latest_file.xlsx"
        s3_client.download_file(bucket_name, latest_file_key, local_path)

        df = pd.read_excel(local_path)

        # MCAの実行
        n_components = min(df.shape) - 1  # use min(rows, columns) - 1 for MCA
        mca = prince.MCA(n_components=n_components)
        mca_result = mca.fit(df)

        # S3への処理後のデータアップロード
        mca_result_df = pd.DataFrame(
            data=mca_result.row_coordinates(df),
            columns=[f'MC{i}' for i in range(1, n_components+1)]
        )
        mca_result_df.to_excel("/tmp/mca_result.xlsx", index=False)
        s3_client.upload_file(
            "/tmp/mca_result.xlsx", bucket_name, "Multiple_correspondence_analysis/mca_result.xlsx")

        # Scree plot (explained variance by component)
        ax = mca.plot_row_coordinates(
            df, figsize=(10, 6), show_row_points=True, show_row_labels=False, show_column_points=True, show_column_labels=True)
        ax.get_figure().savefig("/tmp/mca_scree_plot.png")

        # Cloudinaryへの処理後のイメージアップロード
        cloudinary.config(
            cloud_name=os.environ.get('CLOUDINARY_CLOUD_NAME'),
            api_key=os.environ.get('CLOUDINARY_API_KEY'),
            api_secret=os.environ.get('CLOUDINARY_API_SECRET')
        )

        upload_response = cloudinary.uploader.upload(
            "/tmp/mca_scree_plot.png",
            folder="Multiple_correspondence_analysis",
            use_filename=True,
            unique_filename=False
        )

        # 最後にフロントにurlを返す
        return Response({"graph_url": upload_response['url']}, status=200)
# ===============================================================================================================================


# FUVのエクセルファイルアップロード


# ===============================================================================================================================
@csrf_exempt
def FUVUpload_file(request):
    if request.method == 'POST':
        try:
            print("Access Key:", os.environ.get('AWS_ACCESS_KEY_ID'))
            print("Secret Key:", os.environ.get('AWS_SECRET_ACCESS_KEY'))

            data = json.loads(request.body.decode('utf-8'))

            file_name = f"{uuid.uuid4()}.xlsx"

            workbook = openpyxl.Workbook()
            sheet = workbook.active
            for index, row in enumerate(data):
                for key, value in row.items():
                    if index == 0:
                        header_col = sheet.cell(
                            row=1, column=list(row.keys()).index(key) + 1)
                        header_col.value = key

                    cell = sheet.cell(
                        row=index + 2, column=list(row.keys()).index(key) + 1)
                    cell.value = value

            with open(file_name, 'wb') as f:
                workbook.save(f)

            s3_path = f"fuv/upload/{file_name}"

            s3_client = boto3.client('s3', aws_access_key_id=os.environ.get('AWS_ACCESS_KEY_ID'),
                                     aws_secret_access_key=os.environ.get('AWS_SECRET_ACCESS_KEY'))

            existing_files = list_files(
                s3_client, settings.AWS_STORAGE_BUCKET_NAME, 'fuv/upload/')
            for file_key in existing_files:
                s3_client.delete_object(
                    Bucket=settings.AWS_STORAGE_BUCKET_NAME, Key=file_key)

            s3_client.upload_file(
                file_name, settings.AWS_STORAGE_BUCKET_NAME, s3_path)

            os.remove(file_name)

            file_url = f"https://{settings.AWS_STORAGE_BUCKET_NAME}.s3.amazonaws.com/{s3_path}"

            return JsonResponse({'message': 'Data processed and saved to S3 successfully!', 'file_url': file_url})
        except json.JSONDecodeError:
            return JsonResponse({'message': 'Failed to decode JSON data.'}, status=400)

    return JsonResponse({'message': 'Only POST requests are allowed.'})


@csrf_exempt
def get_concentration_count(request):
    if request.method != 'GET':
        return JsonResponse({'message': 'Only GET requests are allowed.'}, status=405)

    try:
        # S3クライアントの初期化
        s3_client = boto3.client('s3',
                                 aws_access_key_id=os.environ.get(
                                     'AWS_ACCESS_KEY_ID'),
                                 aws_secret_access_key=os.environ.get('AWS_SECRET_ACCESS_KEY'))
        bucket_name = os.environ.get(
            'AWS_STORAGE_BUCKET_NAME')  # 環境変数からバケット名を取得

        # S3バケット内のファイル一覧を取得
        response = s3_client.list_objects_v2(
            Bucket=bucket_name, Prefix='fuv/upload/')
        files = [file['Key'] for file in response.get('Contents', [])]

        if not files:
            return JsonResponse({'message': 'No files found in S3 bucket.'}, status=404)

        # 最新のファイルを取得
        latest_file_key = max(files, key=lambda x: x.split('/')[-1])
        latest_file_obj = s3_client.get_object(
            Bucket=bucket_name, Key=latest_file_key)

        # Excelファイルを読み込む
        df = pd.read_excel(BytesIO(latest_file_obj['Body'].read()))
        # '濃度'列でユニークな値の数を数える（列名はファイルによって変更する必要があるかもしれません）
        concentration_count = df['濃度'].nunique()
        return JsonResponse({'concentration_count': concentration_count})

    except Exception as e:
        logger.error(f'Error getting concentration count: {e}', exc_info=True)
        return JsonResponse({'message': 'Server error while retrieving concentration count.'}, status=500)


@csrf_exempt
def FUVSecondDerivativeUpload(request):
    if request.method == 'POST':
        try:
            file_obj = request.FILES.get('file')

            if not file_obj:
                return JsonResponse({'message': 'File is required.'}, status=400)

            file_name = f"{uuid.uuid4()}.xlsx"
            with open(file_name, 'wb') as f:
                for chunk in file_obj.chunks():
                    f.write(chunk)

            s3_path = f"fuv/second_analysis/{file_name}"

            s3_client = boto3.client('s3',
                                     aws_access_key_id=os.environ.get(
                                         'AWS_ACCESS_KEY_ID'),
                                     aws_secret_access_key=os.environ.get('AWS_SECRET_ACCESS_KEY'))

            existing_files = list_files(
                s3_client, settings.AWS_STORAGE_BUCKET_NAME, 'fuv/second_analysis/')
            for file_key in existing_files:
                s3_client.delete_object(
                    Bucket=settings.AWS_STORAGE_BUCKET_NAME, Key=file_key)

            s3_client.upload_file(
                file_name, settings.AWS_STORAGE_BUCKET_NAME, s3_path)

            os.remove(file_name)

            file_url = f"https://{settings.AWS_STORAGE_BUCKET_NAME}.s3.amazonaws.com/{s3_path}"

            return JsonResponse({'message': 'File uploaded and saved to S3 successfully!', 'file_url': file_url})
        except Exception as e:
            return JsonResponse({'message': str(e)}, status=400)

    return JsonResponse({'message': 'Only POST requests are allowed.'})


@csrf_exempt
def FUVNireUpload_file(request):
    if request.method == 'POST':
        try:
            print("Access Key:", os.environ.get('AWS_ACCESS_KEY_ID'))
            print("Secret Key:", os.environ.get('AWS_SECRET_ACCESS_KEY'))

            data = json.loads(request.body.decode('utf-8'))

            file_name = f"{uuid.uuid4()}.xlsx"

            workbook = openpyxl.Workbook()
            sheet = workbook.active
            for index, row in enumerate(data):
                for key, value in row.items():
                    if index == 0:
                        header_col = sheet.cell(
                            row=1, column=list(row.keys()).index(key) + 1)
                        header_col.value = key

                    cell = sheet.cell(
                        row=index + 2, column=list(row.keys()).index(key) + 1)
                    cell.value = value

            with open(file_name, 'wb') as f:
                workbook.save(f)

            s3_path = f"fuv/nire/{file_name}"

            s3_client = boto3.client('s3', aws_access_key_id=os.environ.get('AWS_ACCESS_KEY_ID'),
                                     aws_secret_access_key=os.environ.get('AWS_SECRET_ACCESS_KEY'))

            existing_files = list_files(
                s3_client, settings.AWS_STORAGE_BUCKET_NAME, 'fuv/nire/')
            for file_key in existing_files:
                s3_client.delete_object(
                    Bucket=settings.AWS_STORAGE_BUCKET_NAME, Key=file_key)

            s3_client.upload_file(
                file_name, settings.AWS_STORAGE_BUCKET_NAME, s3_path)

            os.remove(file_name)

            file_url = f"https://{settings.AWS_STORAGE_BUCKET_NAME}.s3.amazonaws.com/{s3_path}"

            return JsonResponse({'message': 'Data processed and saved to S3 successfully!', 'file_url': file_url})
        except json.JSONDecodeError:
            return JsonResponse({'message': 'Failed to decode JSON data.'}, status=400)

    return JsonResponse({'message': 'Only POST requests are allowed.'})


@csrf_exempt
def smooth_data(data, window_length=5, polyorder=2):
    """
    Smooth the data using Savitzky-Golay filter.

    Parameters:
    - data: The data to be smoothed
    - window_length: The length of the filter window (must be an odd integer)
    - polyorder: The order of the polynomial used to fit the samples

    Returns:
    - smoothed_data: The smoothed data
    """
    return savgol_filter(data, window_length, polyorder)


@csrf_exempt
def kk_transform(incident_angle, n_inf, absorbance_df, fixed_nire_value):
    # Constants
    c = 299792458  # Speed of light in m/s
    incident_angle_rad = np.radians(incident_angle)  # Convert angle to radians

    # Compute frequency from wavelength
    frequency = c * 10**9 / absorbance_df['波長']
    results = []

    # Loop over each concentration column (skipping the wavelength column)
    for column in absorbance_df.columns[1:]:
        absorbance = absorbance_df[column]
        R = 10**(-absorbance)  # Reflectance

        # Initialize arrays for the calculated values
        F = np.zeros(len(frequency))
        nfin = np.zeros(len(frequency))
        kfin = np.zeros(len(frequency))

        # Perform the transformation using the fixed nire values
        for i in range(len(frequency)):
            # Avoid division by zero or negative square roots
            val = fixed_nire_value**2 * \
                np.sin(incident_angle_rad)**2 - n_inf**2
            actn_val = 0 if val < 0 else np.arctan(
                np.sqrt(val) / (fixed_nire_value * np.cos(incident_angle_rad)))
            F[i] = 2 * actn_val

            # Calculate nfin and kfin for each frequency
            r = np.sqrt(R[i]) * np.exp(1j * F[i])
            nfin[i] = fixed_nire_value * np.real(np.sqrt(np.sin(incident_angle_rad)**2 + (
                1 - r) / (1 + r) * np.cos(incident_angle_rad)**2))
            kfin[i] = -fixed_nire_value * np.imag(np.sqrt(np.sin(incident_angle_rad)**2 + (
                1 - r) / (1 + r) * np.cos(incident_angle_rad)**2))

        # Append the results to a list
        result = np.column_stack((absorbance_df['波長'], F, nfin, kfin))
        results.append(result)

    return results


@csrf_exempt
def kk_transformed_spectrum(request):
    # S3クライアントの初期化
    s3_client = boto3.client('s3', aws_access_key_id=os.environ.get('AWS_ACCESS_KEY_ID'),
                             aws_secret_access_key=os.environ.get('AWS_SECRET_ACCESS_KEY'))
    bucket_name = settings.AWS_STORAGE_BUCKET_NAME

    # fuv/uploadから吸光度データを取得
    fuv_files = list_files(s3_client, bucket_name, 'fuv/upload')
    if not fuv_files:
        return HttpResponse('No fuv files found in S3 bucket.')

    # fuv/nire_uploadから屈折率データを取得
    nire_files = list_files(s3_client, bucket_name, 'fuv/nire')
    if not nire_files:
        return HttpResponse('No nire files found in S3 bucket.')

    data = json.loads(request.body.decode("utf-8"))

    # Reactからのデータを抽出
    n_inf = float(data['n_inf'])
    incident_angle = float(data['incident_angle'])

    # 吸光度データをデータフレームとして読み込み
    fuv_obj = s3_client.get_object(Bucket=bucket_name, Key=fuv_files[0])
    df = pd.read_excel(BytesIO(fuv_obj['Body'].read()))

    # 屈折率データをデータフレームとして読み込み
    nire_obj = s3_client.get_object(Bucket=bucket_name, Key=nire_files[0])
    nire_df = pd.read_excel(BytesIO(nire_obj['Body'].read()))

    results = {
        'F': pd.DataFrame(),
        'nfin': pd.DataFrame(),
        'kfin': pd.DataFrame()
    }

    for column in df.columns[1:]:
        if column in nire_df.columns:
            transform_result = kk_transform(incident_angle, n_inf,
                                            df[['波長', column]], nire_df[['波長', column]])
            wl, F, nfin, kfin = transform_result[0].T
            results['F'][column] = F
            results['nfin'][column] = nfin
            results['kfin'][column] = kfin

    # 結果をS3に保存
    for key, dataframe in results.items():
        excel_io = io.BytesIO()
        dataframe.insert(0, "Wavelength", df['波長'])
        dataframe.to_excel(excel_io, index=False)
        excel_io.seek(0)
        s3_client.upload_fileobj(
            excel_io, bucket_name, f'fuv/kk/transformed_{key}.xlsx')

    return HttpResponse('KK transformed data uploaded to S3 successfully.')


def list_files(client, bucket, prefix):
    try:
        response = client.list_objects_v2(Bucket=bucket, Prefix=prefix)
        return [item['Key'] for item in response.get('Contents', [])]
    except Exception as e:
        print(
            f"Error listing files in bucket {bucket} with prefix {prefix}: {e}")
        return []


def kk_download_latest_from_s3(prefixes):
    s3_client = boto3.client('s3',
                             aws_access_key_id=os.environ.get(
                                 'AWS_ACCESS_KEY_ID'),
                             aws_secret_access_key=os.environ.get('AWS_SECRET_ACCESS_KEY'))
    bucket_name = 'newniv-bucket'

    # Create an in-memory output file for the new zip.
    zip_buffer = BytesIO()

    # Open the zip file for writing, and write the S3 files into it.
    with zipfile.ZipFile(zip_buffer, 'a', zipfile.ZIP_DEFLATED, False) as z:
        for prefix in prefixes:
            files = list_files(s3_client, bucket_name, prefix)
            if not files:
                continue  # No files found for this prefix
            latest_file_key = files[-1]
            file_stream = s3_client.get_object(
                Bucket=bucket_name, Key=latest_file_key)['Body']
            z.writestr(latest_file_key.split("/")[-1], file_stream.read())

    # Zip files are written, now we position the stream to the beginning.
    zip_buffer.seek(0)

    # Create the Django response object and set the appropriate headers.
    response = HttpResponse(zip_buffer, content_type='application/zip')
    response['Content-Disposition'] = 'attachment; filename="downloaded_files.zip"'

    return response


def kk_download_all(request):
    print("kk_download_all function called.")
    prefixes = ['fuv/kk/transformed_F.xlsx',
                'fuv/kk/transformed_kfin.xlsx', 'fuv/kk/transformed_nfin.xlsx']
    return kk_download_latest_from_s3(prefixes)


@csrf_exempt
def fuv_second_derivative(request):
    if request.method == 'POST':
        s3_client = boto3.client('s3')
        bucket_name = 'newniv-bucket'

        files = list_files(s3_client, bucket_name, 'fuv/second_analysis/')
        if not files:
            return JsonResponse({'error': 'No files found in S3 bucket.'}, status=400)

        latest_file_key = files[-1]
        local_path = "/tmp/latest_file.xlsx"
        s3_client.download_file(bucket_name, latest_file_key, local_path)

        df = pd.read_excel(local_path)

        # Create a copy of the dataframe to store the second derivative data
        derivative_df = df.copy()

        plt.figure(figsize=(10, 6))
        colors = plt.cm.rainbow(np.linspace(0, 1, len(df.columns) - 1))

        for col, color in zip(df.columns.drop('波長'), colors):
            smoothed_data = ndimage.gaussian_filter1d(df[col], sigma=10)
            y = ndimage.gaussian_filter1d(smoothed_data, sigma=10, order=2)
            derivative_df[col] = y

            plt.plot(df['波長'], y, label=col, color=color)

        plt.title('Second Derivative of NIR Spectrum')
        plt.xlabel('Wavelength (cm-1)')
        plt.ylabel('Second Derivative of Absorbance')
        plt.legend(loc='upper right')

        graph_filename = 'second_derivative_nir_spectrum.png'
        graph_dir = 'static'
        graph_filepath = os.path.join(graph_dir, graph_filename)

        if not os.path.exists(graph_dir):
            os.makedirs(graph_dir)

        plt.savefig(graph_filepath)
        plt.close()

        # Save the second derivative data to S3
        processed_excel_path = os.path.join(
            graph_dir, 'second_derivative_data.xlsx')
        derivative_df.to_excel(processed_excel_path, index=False)
        s3_upload_path = f'fuv/second_derivative/second_derivative_data.xlsx'
        s3_client.upload_file(processed_excel_path,
                              bucket_name, s3_upload_path)

        # Cloudinaryの設定
        cloudinary.config(
            cloud_name=os.environ.get('CLOUDINARY_CLOUD_NAME'),
            api_key=os.environ.get('CLOUDINARY_API_KEY'),
            api_secret=os.environ.get('CLOUDINARY_API_SECRET')
        )

        # Cloudinaryに保存されている古いイメージを削除
        folder_name = 'FuvSecondDerivative'
        stored_images = cloudinary.api.resources(
            type='upload', prefix=f"{folder_name}/", max_results=500)
        for image in stored_images['resources']:
            cloudinary.uploader.destroy(image['public_id'])

        # グラフをCloudinaryのフォルダにアップロード
        upload_response = cloudinary.uploader.upload(
            graph_filepath, folder=folder_name, use_filename=True, unique_filename=False)
        cloudinary_url = upload_response['url']

        response_data = {'graph_url': cloudinary_url}
        return JsonResponse(response_data)

    return JsonResponse({'error': 'Invalid request method'}, status=400)


@csrf_exempt
def fuv_second_derivative_download(request):
    if request.method == 'GET':
        s3_client = boto3.client('s3',
                                 aws_access_key_id=os.environ.get(
                                     'AWS_ACCESS_KEY_ID'),
                                 aws_secret_access_key=os.environ.get('AWS_SECRET_ACCESS_KEY'))
        bucket_name = 'newniv-bucket'
        s3_path = 'fuv/second_derivative/second_derivative_data.xlsx'
        local_path = '/tmp/second_derivative_data.xlsx'

        try:
            # S3からファイルをローカルにダウンロード
            s3_client.download_file(bucket_name, s3_path, local_path)

            # ローカルのファイルをレスポンスとして返す
            with open(local_path, 'rb') as f:
                response = HttpResponse(f.read(
                ), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = 'attachment; filename="second_derivative_data.xlsx"'
                return response

        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)

    return JsonResponse({'error': 'Invalid request method'}, status=400)


@api_view(['POST'])
def find_peak_upload_file(request):
    if request.method == 'POST':
        try:
            # AWSのキー情報をプリント
            print("Access Key:", os.environ.get('AWS_ACCESS_KEY_ID'))
            print("Secret Key:", os.environ.get('AWS_SECRET_ACCESS_KEY'))

            # リクエストからファイルを取得
            file_obj = request.FILES['file']
            file_name = f"{uuid.uuid4()}{file_obj.name}"  # UUIDを付与してファイル名を生成
            s3_path = f"other/find_peak/{file_name}"

            # S3のクライアントを初期化
            s3_client = boto3.client('s3', aws_access_key_id=os.environ.get('AWS_ACCESS_KEY_ID'),
                                     aws_secret_access_key=os.environ.get('AWS_SECRET_ACCESS_KEY'))

            # 同じディレクトリに既存のファイルを削除
            existing_files = list_files(
                s3_client, settings.AWS_STORAGE_BUCKET_NAME, 'other/find_peak/')
            for file_key in existing_files:
                s3_client.delete_object(
                    Bucket=settings.AWS_STORAGE_BUCKET_NAME, Key=file_key)

            # ファイルをS3にアップロード
            s3_client.upload_fileobj(
                file_obj, settings.AWS_STORAGE_BUCKET_NAME, s3_path)

            file_url = f"https://{settings.AWS_STORAGE_BUCKET_NAME}.s3.amazonaws.com/{s3_path}"

            return JsonResponse({'message': 'File uploaded to S3 successfully!', 'file_url': file_url})

        except Exception as e:
            return JsonResponse({'message': f'An error occurred: {e}'}, status=400)

    return JsonResponse({'message': 'Only POST requests are allowed.'})
# ===============================================================================================================================


# ===============================================================================================================================
def list_files(s3_client, bucket_name, prefix):
    response = s3_client.list_objects_v2(Bucket=bucket_name, Prefix=prefix)
    return [content['Key'] for content in response.get('Contents', [])]


def download_and_filter_data(x_start, x_end):
    s3_client = boto3.client('s3', aws_access_key_id=os.environ.get('AWS_ACCESS_KEY_ID'),
                             aws_secret_access_key=os.environ.get('AWS_SECRET_ACCESS_KEY'))
    s3_path = "other/find_peak/"

    # フォルダ内の全ファイルをリストアップ
    objects = s3_client.list_objects_v2(
        Bucket=settings.AWS_STORAGE_BUCKET_NAME, Prefix=s3_path)

    # 最新のファイルを特定
    latest_file = max(objects.get('Contents', []),
                      key=lambda x: x['LastModified'])

    # ファイルのS3上のキーを取得
    file_key = latest_file['Key']

    file_name = "downloaded_file.xlsx"

    with open(file_name, 'wb') as f:
        s3_client.download_fileobj(
            settings.AWS_STORAGE_BUCKET_NAME, file_key, f)

    df = pd.read_excel(file_name)
    os.remove(file_name)

    df_filtered = df[(df['波長'] >= x_start) & (df['波長'] <= x_end)]

    return df_filtered


def find_peak(request):
    x_start = float(request.data.get('x_start'))
    x_end = float(request.data.get('x_end'))

    df_filtered = download_and_filter_data(x_start, x_end)

    # "波長"カラムを除外して、濃度のカラムリストを取得
    concentration_columns = [col for col in df_filtered.columns if col != "波長"]
    peak_data_list = []

    for col in concentration_columns:
        max_intensity = df_filtered[col].max()
        max_wavelength = df_filtered[df_filtered[col]
                                     == max_intensity]['波長'].values[0]
        peak_data = {
            "concentration": col,
            "x": max_wavelength,
            "y": max_intensity
        }
        peak_data_list.append(peak_data)

    return JsonResponse({"data": peak_data_list})


@api_view(['POST'])
def evaluate_peaks_within_range(request):
    x_start = float(request.data.get('x_start'))
    x_end = float(request.data.get('x_end'))

    df_filtered = download_and_filter_data(x_start, x_end)

    # "波長"カラムを除外して、濃度のカラムリストを取得
    concentration_columns = [col for col in df_filtered.columns if col != "波長"]

    peak_data_list = []

    for col in concentration_columns:
        max_intensity = df_filtered[col].max()
        max_wavelength = df_filtered[df_filtered[col]
                                     == max_intensity]['波長'].values[0]

        peak_data = {
            "concentration": col,
            "x": max_wavelength,
            "y": max_intensity
        }
        peak_data_list.append(peak_data)

    # ピーク検出後のデータを新たなデータフレームとして保存
    df_peaks = pd.DataFrame(peak_data_list)
    temp_peak_excel_name = "detected_peak_data.xlsx"
    df_peaks.to_excel(temp_peak_excel_name, index=False)

    # Upload the Excel file with detected peaks to S3
    s3_key_peak_excel = "other/find_peaked/" + temp_peak_excel_name
    s3_client.upload_file(temp_peak_excel_name, bucket_name, s3_key_peak_excel)

    os.remove(temp_peak_excel_name)

    # Prepare the S3 URL for the uploaded Excel file with detected peaks
    s3_url_peak_excel = f"https://{bucket_name}.s3.amazonaws.com/{s3_key_peak_excel}"

    response_data = {
        "excel_data_url": s3_url_peak_excel,
        "peaks": peak_data_list
    }
    return JsonResponse(response_data)


def download_peaks_data(request):
    s3_client = boto3.client('s3',
                             aws_access_key_id=os.environ.get(
                                 'AWS_ACCESS_KEY_ID'),
                             aws_secret_access_key=os.environ.get('AWS_SECRET_ACCESS_KEY'))

    # S3内の'second_derivative/'ディレクトリから最新のファイルを取得
    bucket_name = 'newniv-bucket'
    prefix = 'other/find_peaked/'
    files = list_files(s3_client, bucket_name, prefix)

    if not files:
        return HttpResponse('No files found in S3 bucket under the specified prefix.')

    try:
        # 最新のファイルのキー
        latest_file_key = files[-1]  # 最新のファイルを取得するために[-1]を使用

        # メモリ上のバイナリストリームとしてファイルを取得
        file_stream = s3_client.get_object(
            Bucket=bucket_name, Key=latest_file_key)['Body']

        # クライアントに送信するためのレスポンスを作成
        response = HttpResponse(file_stream.read(),
                                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        # 最新のファイル名をそのまま使用
        response[
            'Content-Disposition'] = f'attachment; filename="{latest_file_key.split("/")[-1]}"'

        return response

    except Exception as e:
        print(e)
        return HttpResponse("An error occurred while downloading the file.")
# ===============================================================================================================================


# ===============================================================================================================================
def smooth_upload_file_to_s3(file):
    s3_client = boto3.client('s3', aws_access_key_id=os.environ.get('AWS_ACCESS_KEY_ID'),
                             aws_secret_access_key=os.environ.get('AWS_SECRET_ACCESS_KEY'))

    folder_path = "other/smooth/"
    s3_path = folder_path + file.name

    # Delete all files in the folder
    objects_to_delete = s3_client.list_objects_v2(
        Bucket=settings.AWS_STORAGE_BUCKET_NAME, Prefix=folder_path)
    if 'Contents' in objects_to_delete:
        for obj in objects_to_delete['Contents']:
            s3_client.delete_object(
                Bucket=settings.AWS_STORAGE_BUCKET_NAME, Key=obj['Key'])

    s3_client.upload_fileobj(file, settings.AWS_STORAGE_BUCKET_NAME, s3_path)
    return s3_path


@api_view(['POST'])
def upload_file_for_smoothing(request):
    if 'file' not in request.FILES:
        return JsonResponse({"error": "File not provided"}, status=400)

    file = request.FILES['file']
    s3_path = smooth_upload_file_to_s3(file)

    return JsonResponse({"message": "File uploaded successfully", "s3_path": s3_path})


class SmoothingData(APIView):

    def post(self, request):
        window_size = int(request.data.get('window_size', 5))  # デフォルト値は5
        polynomial_order = int(request.data.get(
            'polynomial_order', 3))  # デフォルト値は3

        s3_client = boto3.client('s3', aws_access_key_id=os.environ.get('AWS_ACCESS_KEY_ID'),
                                 aws_secret_access_key=os.environ.get('AWS_SECRET_ACCESS_KEY'))
        bucket_name = settings.AWS_STORAGE_BUCKET_NAME

        # S3 path prefix for smoothing data
        s3_prefix = 'other/smooth/'
        files = list_files(s3_client, bucket_name, s3_prefix)
        if not files:
            return JsonResponse({"error": "No files found in S3 bucket."}, status=400)

        latest_file_key = files[-1]
        local_path = "/tmp/latest_file_for_smoothing.xlsx"
        s3_client.download_file(bucket_name, latest_file_key, local_path)

        df = pd.read_excel(local_path)

        # Apply Savitzky-Golay filter for smoothing
        for column in df.columns:
            if df[column].dtype == "float64":
                df[column] = savgol_filter(
                    df[column], window_size, polynomial_order)

        # Save smoothed data to a new Excel file
        output_file_name = "/tmp/smoothed_data.xlsx"
        df.to_excel(output_file_name, index=False)

        # Upload smoothed file to S3
        smoothed_s3_path = "other/smoothed_files/" + \
            os.path.basename(output_file_name)
        s3_client.upload_file(output_file_name, bucket_name, smoothed_s3_path)

        return JsonResponse({"message": "Data smoothed successfully", "smoothed_s3_path": smoothed_s3_path})


def download_smoothed_data(request):
    s3_client = boto3.client('s3',
                             aws_access_key_id=os.environ.get(
                                 'AWS_ACCESS_KEY_ID'),
                             aws_secret_access_key=os.environ.get('AWS_SECRET_ACCESS_KEY'))

    # S3内の'second_derivative/'ディレクトリから最新のファイルを取得
    bucket_name = 'newniv-bucket'
    prefix = 'other/smoothed_files/'
    files = list_files(s3_client, bucket_name, prefix)

    if not files:
        return HttpResponse('No files found in S3 bucket under the specified prefix.')

    try:
        # 最新のファイルのキー
        latest_file_key = files[-1]  # 最新のファイルを取得するために[-1]を使用

        # メモリ上のバイナリストリームとしてファイルを取得
        file_stream = s3_client.get_object(
            Bucket=bucket_name, Key=latest_file_key)['Body']

        # クライアントに送信するためのレスポンスを作成
        response = HttpResponse(file_stream.read(),
                                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        # 最新のファイル名をそのまま使用
        response[
            'Content-Disposition'] = f'attachment; filename="{latest_file_key.split("/")[-1]}"'

        return response

    except Exception as e:
        print(e)
        return HttpResponse("An error occurred while downloading the file.")
# ===============================================================================================================================
